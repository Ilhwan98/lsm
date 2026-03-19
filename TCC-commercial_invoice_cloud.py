import os
import io
import json
import time
from datetime import datetime, timedelta
import pandas as pd
import pdfplumber
from httplib2 import Http
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from google.oauth2 import service_account
import openpyxl


SCOPES = ["https://www.googleapis.com/auth/drive"]
SERVICE_ACCOUNT_FILE = "service_account.json"

SOURCE_FOLDER_ID = os.getenv("GOOGLE_DRIVE_SOURCE_FOLDER_ID")
TARGET_FOLDER_ID = os.getenv("GOOGLE_DRIVE_TARGET_FOLDER_ID")
ERROR_URL = os.getenv("https://chat.googleapis.com/v1/spaces/AAAAdp2-7rc/messages?key=AIzaSyDdI0hCZtE6vySjMm-WEfRq3CPzqKqqsHI&token=BVhGARAAbgjqHRJVQz-NIoIPO_IgcIYcdATwu1v9bEw")

today = datetime.now()
today_date = today.strftime("%m%d%y")
current_date = today

BASE_DIR = os.getcwd()
DOWNLOAD_DIR = os.path.join(BASE_DIR, "downloads")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

os.makedirs(DOWNLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

def get_drive_service():
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE,
        scopes=SCOPES
    )
    return build("drive", "v3", credentials=creds)


def list_drive_files(service, query):
    files = []
    page_token = None

    while True:
        response = service.files().list(
            q=query,
            fields="nextPageToken, files(id, name, mimeType)",
            pageToken=page_token,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True
        ).execute()

        files.extend(response.get("files", []))
        page_token = response.get("nextPageToken")

        if not page_token:
            break

    return files

def find_today_billing_folder(service, parent_folder_id, today_date):
    query = (
        f"'{parent_folder_id}' in parents and trashed=false and "
        f"mimeType='application/vnd.google-apps.folder' and "
        f"name contains '{today_date} Billing'"
    )

    folders = list_drive_files(service, query)

    if not folders:
        return None

    for folder in folders:
        if folder["name"].startswith(f"{today_date} Billing"):
            return folder

    return folders[0]

def list_files_in_folder(service, folder_id):
    query = f"'{folder_id}' in parents and trashed=false"
    return list_drive_files(service, query)

def download_drive_file(service, file_id, destination):
    request = service.files().get_media(fileId=file_id, supportsAllDrives=True)

    with open(destination, "wb") as f:
        downloader = MediaIoBaseDownload(f, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()


def download_ent_pdfs(service, billing_folder_id):
    print("📥 Downloading PDFs...")

    sheet_names = []

    items = list_drive_files(service, f"'{billing_folder_id}' in parents")

    for item in items:
        if item["mimeType"] == "application/vnd.google-apps.folder" and item["name"].startswith("SEE"):
            print("📂 SEE folder:", item["name"])
            sheet_names.append(item["name"])

            files = list_drive_files(service, f"'{item['id']}' in parents")

            for f in files:
                if f["name"].endswith(".ENT.pdf"):
                    path = os.path.join(DOWNLOAD_DIR, f"{item['name']}_{f['name']}")
                    print("⬇️ Downloading:", f["name"])
                    download_drive_file(service, f["id"], path)

    return sheet_names


def extract_pdf_to_csv(pdf_path, csv_path):
    tables = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                tables.extend(table)

    if not tables:
        print("⚠️ No tables found:", pdf_path)
        return False

    max_cols = max(len(row) for row in tables)
    rows = [row + [""] * (max_cols - len(row)) for row in tables]

    df = pd.DataFrame(rows[1:], columns=rows[0])
    df.to_csv(csv_path, index=False)

    return True

def process_csv_to_excel(sheet, csv_path):
    df = pd.read_csv(csv_path)

    for r, row in df.iterrows():
        for c, value in enumerate(row, start=1):
            sheet.cell(row=r+1, column=c, value=str(value))

def read_file(path, csv_filename, count, sheet, ost, newfile_name, error_url=None):
    count_start = count
    info = []
    data = []
    data_set = {}
    data_fin = []
    title = ""

    try:
        read_csv_path = os.path.join(path, csv_filename)
        df = pd.read_csv(read_csv_path)

        for index, row in df.iterrows():
            for col in df.columns:
                value = row[col]

                if isinstance(value, str) and "7 Release Date" in value:
                    extracted_info = value.split("7 Release Date")[1].strip()
                    info.append(("Release Date: ", extracted_info))

                if isinstance(value, str) and "12 Importer Name, Address and Telephone No." in value:
                    extracted_info1 = value.split("12 Importer Name, Address and Telephone No.")[1].strip()
                    formatted_info1 = extracted_info1.replace("\n", " ")
                    info.append(("Importer Info:", formatted_info1))

                if isinstance(value, str) and "36 Vendor - Name, Address and Telephone No." in value:
                    extracted_info = value.split("36 Vendor - Name, Address and Telephone No.")[1].strip()
                    formatted_info = " ".join(line.strip() for line in extracted_info.splitlines() if line.strip())
                    info.append(("Vendor Info: ", formatted_info))

                if isinstance(value, str) and "15 Cargo Control No." in value:
                    extracted_info = value.split("15 Cargo Control No.")[1].strip()
                    info.append(("Cargo Control No.: ", extracted_info))
                    if "A1X-" in extracted_info:
                        title = extracted_info.split("A1X-")[1].strip()
                    else:
                        title = extracted_info[-10:].strip()

                if isinstance(value, str) and "19 Original Transaction No." in value:
                    extracted_info = value.split("19 Original Transaction No.")[1].strip()
                    info.append(("Original Transaction No.: ", extracted_info))

                if isinstance(value, str) and "77 Exchange Rate" in value:
                    extracted_info = value.split("77 Exchange Rate")[1].strip()
                    info.append(("Exchange Rate: ", extracted_info))

                if isinstance(value, str) and "39 Invoice Value" in value:
                    extracted_info = "$" + value.split("39 Invoice Value")[1].strip()
                    info.append(("Invoice Value (USD): ", extracted_info))

                if isinstance(value, str) and "130 TOTAL DUTIES & TAXES " in value:
                    extracted_info = "$" + value.split("130 TOTAL DUTIES & TAXES ")[1].strip()
                    info.append(("TOTAL DUTIES & TAXES (CAD): ", extracted_info))

                if isinstance(value, str) and "56 Line No." in value:
                    num = value.split("56 Line No.")[1].strip()
                    if num != "":
                        if index + 4 >= len(df):
                            continue

                        now_row = df.iloc[index]
                        next_row = df.iloc[index + 1]
                        two_row = df.iloc[index + 2]
                        three_row = df.iloc[index + 3]
                        four_row = df.iloc[index + 4]

                        country = ""
                        hsc = ""
                        exch_rate = 0.0
                        val_currency = 0.0
                        customs = 0.0
                        gst = 0.0
                        duty = 0.0
                        val_duty = 0.0
                        val_tax = 0.0

                        for next_col in df.columns:
                            value_cell = now_row[next_col]
                            value1_cell = next_row[next_col]
                            value2_cell = two_row[next_col]
                            value3_cell = three_row[next_col]
                            value4_cell = four_row[next_col]

                            if isinstance(value1_cell, str) and "65 COO" in value1_cell:
                                country = str(value1_cell.split("65 COO")[1].strip().replace(",", ""))

                            if isinstance(value_cell, str) and "58 Classification No." in value_cell:
                                hsc = str(value_cell.split("58 Classification No.")[1].strip().replace(",", ""))

                            if isinstance(value2_cell, str) and "77 Exchange Rate" in value2_cell:
                                exch_rate = float(value2_cell.split("77 Exchange Rate")[1].strip().replace(",", ""))

                            if isinstance(value2_cell, str) and "75 Value for Currency Conversion" in value2_cell:
                                val_currency = float(value2_cell.split("75 Value for Currency Conversion")[1].strip().replace(",", ""))

                            if isinstance(value3_cell, str) and "82 Customs Duty" in value3_cell:
                                customs = float(value3_cell.split("82 Customs Duty")[1].strip().replace(",", ""))

                            if isinstance(value3_cell, str) and "90 GST" in value3_cell:
                                gst = float(value3_cell.split("90 GST")[1].strip().replace(",", ""))

                            if isinstance(value4_cell, str) and "100 Commodity Duty & Tax" in value4_cell:
                                duty = float(value4_cell.split("100 Commodity Duty & Tax")[1].strip().replace(",", ""))

                            if isinstance(value2_cell, str) and "78 Value for Duty" in value2_cell:
                                val_duty = float(value2_cell.split("78 Value for Duty")[1].strip().replace(",", ""))

                            if isinstance(value3_cell, str) and "89 Value for Tax" in value3_cell:
                                val_tax = float(value3_cell.split("89 Value for Tax")[1].strip().replace(",", ""))

                        duty_rate = "0.00%"
                        if val_duty != 0:
                            duty_rate = "{:.2f}%".format((customs / val_duty) * 100)

                        data.append((country, hsc, exch_rate, val_duty, val_tax, val_currency, customs, gst, duty, duty_rate))
                        info.append(("Exchange Rate: ", str(exch_rate) + "%"))

        for country, hsc, exch_rate, val_duty, val_tax, val_currency, customs, gst, duty, duty_rate in data:
            key = (country, hsc)
            if key not in data_set:
                data_set[key] = [exch_rate, val_duty, val_tax, val_currency, customs, gst, duty, duty_rate]
            else:
                data_set[key][0] = exch_rate
                data_set[key][1] += val_duty
                data_set[key][2] += val_tax
                data_set[key][3] += val_currency
                data_set[key][4] += customs
                data_set[key][5] += gst
                data_set[key][6] += duty
                data_set[key][7] = duty_rate

        for (country, hsc), (exch_rate, val_duty, val_tax, val_currency, customs, gst, duty, duty_rate) in data_set.items():
            data_fin.append(("Country of Origin: ", country))
            data_fin.append(("HS Code: ", hsc))
            data_fin.append(("Value for Currency Conversion (USD): ", "${:.2f}".format(val_currency)))
            data_fin.append(("Duty Rate: ", duty_rate))
            data_fin.append(("Commodity Duty & Tax: ", "${:.2f}".format(duty)))

        border_bot = Border(bottom=Side(style="thin"))
        border_2bot = Border(bottom=Side(style="double"))
        border_right = Border(right=Side(style="thin"))
        border_both = Border(bottom=Side(style="thin"), right=Side(style="thin"))
        border_2both = Border(bottom=Side(style="double"), right=Side(style="thin"))
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_font = Font(color="FF0000")

        while len(info) < 8:
            info.append(("", ""))

        reordered_row_data = [info[0], info[4], info[1], info[2], info[6], info[5], info[3], info[7]]

        sheet.merge_cells(start_row=count, start_column=1, end_row=count, end_column=2)
        merged_cell = sheet.cell(row=count, column=1)
        merged_cell.value = title
        merged_cell.border = border_both
        merged_cell.font = Font(bold=True)
        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
        merged_cell.fill = yellow_fill
        sheet.cell(row=count, column=2).border = border_both

        trans_row = count + 1
        for row_idx, row_data in enumerate(reordered_row_data, start=count + 1):
            row_count = count + 8
            trans_row = row_count

            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                if row_idx == count + 2 and col_idx == 2:
                    cell.font = red_font
                if col_idx == 2:
                    cell.border = border_right

            sheet.cell(row=row_count, column=1).border = border_2bot
            sheet.cell(row=row_count, column=2).border = border_2both

        for row_idx, row_data in enumerate(data_fin, start=count + 9):
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

        end_row = row_idx if data_fin else count + 9
        test = count + 8

        for current_row in range(test + 1, end_row, 1):
            for col_idx in range(1, 3):
                cell = sheet.cell(row=current_row, column=col_idx)
                if col_idx == 2:
                    cell.border = border_right

        for current_row in range(test + 5, end_row, 5):
            sheet.cell(row=current_row, column=1).border = border_bot
            sheet.cell(row=current_row + 5, column=1).border = border_bot
            sheet.cell(row=current_row, column=2).border = border_both
            sheet.cell(row=current_row + 5, column=2).border = border_both

        sheet.cell(row=end_row + 1, column=1).fill = black_fill
        sheet.cell(row=end_row + 1, column=2).fill = black_fill
        sheet.cell(row=end_row, column=2).border = border_both

        count = end_row + 2

        for column in sheet.columns:
            column_cells = [cell for cell in column]
            try:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
            except ValueError:
                max_length = 0
            adjusted_width = max_length + 2
            if len(column_cells) > 0:
                sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        full_path = os.path.join(OUTPUT_DIR, newfile_name)
        ost.save(full_path)
        return count

    except Exception as e:
        return count_start

def upload_drive_file(service, file_path, filename):
    print("📤 Uploading file...")

    media = MediaFileUpload(file_path, resumable=True)

    file_metadata = {
        "name": filename,
        "parents": [TARGET_FOLDER_ID]
    }

    file = service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id, name",
        supportsAllDrives=True
    ).execute()

    print("✅ Uploaded:", file["name"])
    return file

def run():
    print("🚀 SCRIPT STARTED")

    service = get_drive_service()

    billing_folder = find_today_billing_folder(service, SOURCE_FOLDER_ID, today_date)
    if not billing_folder:
        raise Exception("Billing folder not found")

    print("📁 Found:", billing_folder["name"])

    sheet_names = download_ent_pdfs(service, billing_folder["id"])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    newfile_name = f"{today_date}_invoice.xlsx"

    for name in sheet_names:
        count = 1

        if name in wb.sheetnames:
            sheet = wb[name]
        else:
            sheet = wb.create_sheet(name)

        for file in os.listdir(DOWNLOAD_DIR):
            if file.startswith(name) and file.endswith(".pdf"):
                pdf_path = os.path.join(DOWNLOAD_DIR, file)
                csv_path = pdf_path.replace(".pdf", ".csv")
                csv_filename = os.path.basename(csv_path)

                if extract_pdf_to_csv(pdf_path, csv_path):
                    print(f"Processing {file}")
                    count = read_file(
                        DOWNLOAD_DIR,
                        csv_filename,
                        count,
                        sheet,
                        wb,
                        newfile_name,
                        error_url=None
                    )

    output_file = os.path.join(OUTPUT_DIR, newfile_name)

    if not wb.sheetnames:
        sheet = wb.create_sheet("NoData")
        sheet["A1"] = "No data processed"

    wb.save(output_file)

    upload_drive_file(service, output_file, os.path.basename(output_file))

    print("🎉 DONE")


if __name__ == "__main__":
    run()